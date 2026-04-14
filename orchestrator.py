"""
Loan Automator — single entry point for scheduled tasks, webhooks, or local runs.

Sequence:
  1) Placement fee patcher (MA API)
  2) Master Tracker enricher (MA API → xlsx, preserves Origination Fee when MA is empty)
  3) Backfill Origination Fee from patcher workbook for rows still empty/zero
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import config

log_file = config.LOGS_DIR / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("orchestrator")


def _ensure_utf8_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def verify_environment(*, ping_api: bool = False) -> int:
    """Return 0 if checks pass, non-zero otherwise."""
    ok = True
    log.info("--- verify: paths ---")
    for label, path in (
        ("EXCEL_PATCHER_PATH", config.EXCEL_PATCHER_PATH),
        ("EXCEL_TRACKER_PATH", config.EXCEL_TRACKER_PATH),
    ):
        p = Path(path)
        if not p.is_file():
            log.error("  %s not found: %s", label, path)
            ok = False
        else:
            log.info("  %s ok (%s)", label, p)

    log.info("--- verify: MA credentials ---")
    if not config.MA_LENDER_ID or not config.MA_API_KEY:
        log.error("  MA_LENDER_ID / MA_API_KEY missing (set in .env or DOTENV_PATH).")
        ok = False
    else:
        log.info("  MA_LENDER_ID present; MA_API_KEY present")

    log.info("--- verify: workbook structure ---")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(config.EXCEL_PATCHER_PATH, read_only=False, data_only=True)
        tbl_found = False
        for sheet in wb.worksheets:
            if config.PATCHER_TABLE_NAME in sheet.tables:
                log.info("  Patcher: table %r on sheet %r", config.PATCHER_TABLE_NAME, sheet.title)
                tbl_found = True
                break
        wb.close()
        if not tbl_found:
            log.error("  Patcher: table %r not found", config.PATCHER_TABLE_NAME)
            ok = False
    except Exception as exc:
        log.exception("  Patcher workbook read failed: %s", exc)
        ok = False

    try:
        from openpyxl import load_workbook

        wb = load_workbook(config.EXCEL_TRACKER_PATH, read_only=False, data_only=True)
        if config.ENRICHER_SHEET_NAME not in wb.sheetnames:
            log.error("  Tracker: missing sheet %r", config.ENRICHER_SHEET_NAME)
            ok = False
        else:
            log.info("  Tracker: sheet %r present", config.ENRICHER_SHEET_NAME)
        wb.close()
    except Exception as exc:
        log.exception("  Tracker workbook read failed: %s", exc)
        ok = False

    if ping_api and ok:
        import os

        import requests

        from steps.placement_patcher import api_auth_token

        loan_raw = os.getenv("VERIFY_LOAN_ID", "").strip()
        if not loan_raw:
            log.warning("  ping_api: set VERIFY_LOAN_ID in env to test loans/get")
        else:
            loan_id = int(loan_raw)
            url = f"{config.MA_API_BASE_V1}/loans/get"
            r = requests.post(
                url,
                headers={
                    "Content-Type": "application/json",
                    "ACCOUNT-ID": config.MA_LENDER_ID,
                    "API-AUTH": api_auth_token("loans/get"),
                },
                json={"loan_id": loan_id},
                timeout=config.TIMEOUT_S,
            )
            log.info("  ping_api: POST loans/get status=%s loan_id=%s", r.status_code, loan_id)
            if r.status_code != 200:
                log.error("  ping_api body (truncated): %s", (r.text or "")[:400])
                ok = False

    if ok:
        log.info("--- verify: ALL OK ---")
        return 0
    log.error("--- verify: FAILED ---")
    return 1


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Loan Automator pipeline")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="No loans/update; enricher/backfill do not save workbooks (GETs still run).",
    )
    parser.add_argument(
        "--verify",
        action="store_true",
        help="Check paths, tables, credentials; exit without running the pipeline.",
    )
    parser.add_argument(
        "--verify-api",
        action="store_true",
        help="With --verify, also POST loans/get for VERIFY_LOAN_ID.",
    )
    parser.add_argument(
        "--step",
        choices=("all", "placement", "enrich", "backfill"),
        default="all",
        help="Run a single step (default: all).",
    )
    parser.add_argument(
        "--placement-last-n",
        type=int,
        default=None,
        metavar="N",
        help="Pilot: only the last N rows of the patcher worklist (table order).",
    )
    parser.add_argument(
        "--placement-first-n",
        type=int,
        default=None,
        metavar="N",
        help="Pilot: only the first N rows of the patcher worklist (table order).",
    )
    parser.add_argument(
        "--enrich-lookback-days",
        type=int,
        default=None,
        metavar="N",
        help="Override enricher window to last N calendar days (UTC), ignoring .env mode for this run.",
    )
    args = parser.parse_args(argv)

    _ensure_utf8_stdio()

    log.info("Log file: %s", log_file)
    log.info(
        "Secrets: DOTENV_PATH=%s (project .env used if unset)",
        os.getenv("DOTENV_PATH") or "(unset)",
    )
    log.info(
        "Env: EXCEL_PATCHER_PATH=%s EXCEL_TRACKER_PATH=%s dry_run=%s step=%s",
        config.EXCEL_PATCHER_PATH,
        config.EXCEL_TRACKER_PATH,
        args.dry_run,
        args.step,
    )

    if args.verify:
        return verify_environment(ping_api=args.verify_api)

    if args.placement_last_n is not None and args.placement_first_n is not None:
        log.error("Use only one of --placement-last-n or --placement-first-n")
        return 2

    if args.step in ("all", "placement", "enrich"):
        config.require_ma_credentials()

    placement_results: list = []
    enrich_summary: dict = {}

    if args.step in ("all", "placement"):
        log.info("STEP 1: Placement fee patcher")
        from steps.placement_patcher import run_patcher

        placement_results = run_patcher(
            dry_run=args.dry_run,
            placement_first_n=args.placement_first_n,
            placement_last_n=args.placement_last_n,
        )
        patched = sum(1 for r in placement_results if r["action"] == "UPDATED")
        skipped = sum(1 for r in placement_results if r["action"] == "SKIPPED_ALREADY_SET")
        errors = sum(1 for r in placement_results if r["action"] == "ERROR")
        get_errs = sum(1 for r in placement_results if r.get("action") == "ERROR_LOANS_GET")
        old_core = sum(1 for r in placement_results if r["action"] == "NOT_SUPPORTED_OLD_CORE")
        dry_would = sum(
            1
            for r in placement_results
            if r.get("action")
            in ("DRY_RUN_WOULD_ATTEMPT_PATCH", "WOULD_PATCH", "WOULD_REVIEW")
        )
        log.info(
            "Step1 summary: n=%s patched=%s skipped=%s errors=%s loans_get_errors=%s "
            "old_core=%s dry_would_attempt=%s",
            len(placement_results),
            patched,
            skipped,
            errors,
            get_errs,
            old_core,
            dry_would,
        )

    if args.step in ("all", "enrich"):
        log.info("STEP 2: Points enricher")
        from steps.points_enricher import run_enricher

        enrich_summary = run_enricher(
            dry_run=args.dry_run,
            lookback_days_override=args.enrich_lookback_days,
        )
        log.info("Step2 summary: %s", enrich_summary)

    if args.step in ("all", "backfill"):
        if args.dry_run and args.step == "all":
            log.info(
                "STEP 3: skipped separate disk backfill in full dry-run "
                "(preview already written after step2 on in-memory merge)."
            )
        else:
            log.info("STEP 3: Placement fee backfill (Excel -> Master)")
            from steps.placement_fee_backfill import run_backfill

            bf_summary = run_backfill(dry_run=args.dry_run)
            log.info("Step3 summary: %s", bf_summary)

    if args.dry_run and args.step == "all":
        from steps.dry_run_report import dry_run_stamp, write_dry_run_aggregate_csv

        agg_stamp = dry_run_stamp()
        agg_path = write_dry_run_aggregate_csv(
            agg_stamp, placement_results, enrich_summary
        )
        log.info("Dry-run aggregate CSV -> %s", agg_path)

    log.info("RUN COMPLETE")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
