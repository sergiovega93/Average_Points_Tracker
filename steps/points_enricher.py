"""
Master Tracker enricher — MA loans/get, merge into workbook.
Preserves Origination Fee when MA returns empty/zero (see config.ORIGINATION_FEE_COLUMN).
"""
from __future__ import annotations

import hashlib
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm

import config


def _api_auth_token() -> str:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d-%H")
    raw = f"{config.MA_LENDER_ID}-{config.MA_API_KEY}-loans/get-{ts}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _headers() -> dict:
    return {
        "ACCOUNT-ID": config.MA_LENDER_ID,
        "API-AUTH": _api_auth_token(),
        "Content-Type": "application/json",
    }


FIELD_MAP = {
    "result.property.address.street": "Street",
    "result.property.address.city": "City",
    "result.property.address.prov": "State",
    "result.property.address.zip": "Zip Code",
    "result.mortgage.interest": "Interest Rate",
    "result.mortgage.funding_date": "Funding Date",
    "result.mortgage.term": "Term Length",
    "result.mortgage.total": "Total Loan Amount",
    "result.mortgage.fees.placement_fee": "Origination Fee",
    "result.repair_cost": "Rehab Budget",
    "result.property.purchase_price": "Purchase Price",
    "result.creation_date": "Creation Date",
    "result.mortgage.type": "Loan Type",
    "result.status": "Status Pipeline",
    "result.custom_fields.macf_ltc_uw_2.response": "LTC UW",
    "result.custom_fields.macf_ltv_uw.response": "LTV UW",
}

ENRICHMENT_ORDER = list(FIELD_MAP.values()) + [
    "Sales Rep",
    "Office Rep",
    "Terms Sent Date",
    "Terms Accepted Date",
    "Title Pending Date",
    "Docs Pending Date",
    "Review Date",
    "Ready to Send Date",
    "Ready to Fund Date",
    "Funded Date",
    "Days: Terms Sent Date → Terms Accepted Date",
    "Days: Terms Accepted Date → Title Pending Date",
    "Days: Title Pending Date → Docs Pending Date",
    "Days: Docs Pending Date → Review Date",
    "Days: Review Date → Ready to Send Date",
    "Days: Ready to Send Date → Ready to Fund Date",
    "Days: Ready to Fund Date → Funded Date",
]


def deep_get(dictionary, keys, default=None):
    for key in keys.split("."):
        if "[" in key and "]" in key:
            base, idx_part = key.split("[", 1)
            index = int(idx_part.replace("]", ""))
            dictionary = dictionary.get(base, [])
            dictionary = dictionary[index] if len(dictionary) > index else {}
        else:
            dictionary = dictionary.get(key, {}) if isinstance(dictionary, dict) else {}
        if dictionary in (None, {}, []):
            return default
    return dictionary


def get_role_user(data, role_name: str) -> str:
    for role in data.get("result", {}).get("roles", []):
        if role.get("name") == role_name and role.get("users"):
            return role["users"][0].get("name", "") or ""
    return ""


def extract_status_dates(data) -> dict:
    milestone_map = {
        "Terms Sent": "Terms Sent Date",
        "Terms Accepted": "Terms Accepted Date",
        "Title Pending": "Title Pending Date",
        "Docs Pending": "Docs Pending Date",
        "Review": "Review Date",
        "Ready to Send": "Ready to Send Date",
        "Ready to Fund": "Ready to Fund Date",
        "funded loan": "Funded Date",
    }
    result = {v: "" for v in milestone_map.values()}
    timestamps: dict = {}

    for note in data.get("result", {}).get("notes", []):
        text = note.get("note", "")
        ts = note.get("date", 0)
        date_str = (
            datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d") if ts else ""
        )

        if text.strip() == "funded loan":
            result["Funded Date"] = date_str
            timestamps["Funded Date"] = ts
        elif text.strip() == "Auto-imported from API":
            result["Terms Sent Date"] = date_str
            timestamps["Terms Sent Date"] = ts
        elif text.startswith("changed status to "):
            full = text.replace("changed status to ", "").strip()
            for base, col in milestone_map.items():
                if base in full:
                    result[col] = date_str
                    timestamps[col] = ts
                    break

    pairs = [
        ("Terms Sent Date", "Terms Accepted Date"),
        ("Terms Accepted Date", "Title Pending Date"),
        ("Title Pending Date", "Docs Pending Date"),
        ("Docs Pending Date", "Review Date"),
        ("Review Date", "Ready to Send Date"),
        ("Ready to Send Date", "Ready to Fund Date"),
        ("Ready to Fund Date", "Funded Date"),
    ]
    for start, end in pairs:
        key = f"Days: {start} → {end}"
        if start in timestamps and end in timestamps:
            result[key] = (timestamps[end] - timestamps[start]) // 86400
        else:
            result[key] = ""

    return result


def _fee_is_missing_or_zero(val) -> bool:
    if val is None or val == "":
        return True
    try:
        return abs(float(val)) <= config.NONZERO_TOL
    except (TypeError, ValueError):
        return True


def _utc_now_naive() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _parse_iso_utc_to_naive(s: str) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _enricher_cutoff_naive() -> datetime:
    """
    Rows with Creation Date >= cutoff (or Creation Date NaN) are enriched.
    """
    now_n = _utc_now_naive()
    mode = (config.ENRICHER_LOOKBACK_MODE or "days").strip().lower()
    if mode == "since_last_run":
        p = Path(config.ENRICHER_LAST_RUN_FILE)
        if p.is_file():
            try:
                last = _parse_iso_utc_to_naive(p.read_text(encoding="utf-8"))
                if last is not None:
                    overlap = timedelta(hours=config.ENRICHER_LOOKBACK_OVERLAP_HOURS)
                    return last - overlap
            except OSError:
                pass
        return now_n - timedelta(days=config.ENRICHER_LOOKBACK_DAYS)
    return now_n - timedelta(days=config.ENRICHER_LOOKBACK_DAYS)


def _write_last_enricher_run_marker() -> None:
    p = Path(config.ENRICHER_LAST_RUN_FILE)
    p.parent.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).isoformat()
    p.write_text(stamp + "\n", encoding="utf-8")


def run_enricher(*, dry_run: bool = False) -> dict:
    """
    Enrich Master Tracker from MA. Does not save workbook if dry_run.
    Returns a small summary dict (rows written, preserved fees count).
    """
    cutoff = _enricher_cutoff_naive()
    mode = (config.ENRICHER_LOOKBACK_MODE or "days").strip().lower()
    print(
        f"  Enricher window: mode={mode!r} cutoff_utc_naive={cutoff:%Y-%m-%d %H:%M} "
        f"lookback_days={config.ENRICHER_LOOKBACK_DAYS} "
        f"overlap_h={config.ENRICHER_LOOKBACK_OVERLAP_HOURS}"
    )
    if mode == "since_last_run":
        p = Path(config.ENRICHER_LAST_RUN_FILE)
        print(f"  Last-run marker file: {p} (exists={p.is_file()})")

    df = pd.read_excel(
        config.EXCEL_TRACKER_PATH,
        sheet_name=config.ENRICHER_SHEET_NAME,
        engine="openpyxl",
    )
    if "Creation Date" in df.columns:
        df["Creation Date"] = pd.to_datetime(df["Creation Date"], errors="coerce")
        mask = df["Creation Date"].isna() | (df["Creation Date"] >= cutoff)
        loan_ids = df.loc[mask, config.API_LOAN_ID_COLUMN].dropna().astype(int)
    else:
        loan_ids = df[config.API_LOAN_ID_COLUMN].dropna().astype(int)

    print(f"  Enriching {len(loan_ids)} loan(s)...")
    enriched_rows = []
    t0 = time.time()
    session = requests.Session()

    for loan_id in tqdm(loan_ids, desc="  MA API", unit="loan"):
        try:
            r = session.post(
                "https://app.mortgageautomator.com/api/loans/get",
                headers=_headers(),
                json={"loan_id": int(loan_id)},
                timeout=config.TIMEOUT_S,
            )
            data = r.json()
            row = {config.API_LOAN_ID_COLUMN: int(loan_id)}
            for key, label in FIELD_MAP.items():
                val = deep_get(data, key)
                if "date" in key and isinstance(val, int):
                    val = datetime.fromtimestamp(
                        val, tz=timezone.utc
                    ).strftime("%Y-%m-%d")
                row[label] = "" if val in (None, {}, []) else val
            row["Sales Rep"] = get_role_user(data, "Sales")
            row["Office Rep"] = get_role_user(data, "Office Rep")
            row.update(extract_status_dates(data))
            enriched_rows.append(row)
        except Exception as e:
            enriched_rows.append({config.API_LOAN_ID_COLUMN: int(loan_id), "Error": str(e)})

        time.sleep(config.SLEEP_BETWEEN_CALLS)

    print(f"  Fetched in {time.time() - t0:.1f}s — merging into sheet...")
    df_enriched = pd.DataFrame(enriched_rows)
    wb = load_workbook(config.EXCEL_TRACKER_PATH)
    ws = wb[config.ENRICHER_SHEET_NAME]

    header_row = [cell.value for cell in ws[1]]
    static_cols = 3
    while (
        static_cols < len(header_row)
        and header_row[static_cols]
        and header_row[static_cols] not in ENRICHMENT_ORDER
    ):
        static_cols += 1
    col_start = static_cols + 1
    api_id_list = [cell.value for cell in ws["B"][1:]]

    orig_off = ENRICHMENT_ORDER.index(config.ORIGINATION_FEE_COLUMN)
    preserved = 0
    per_loan_rows: list[dict] = []
    n_orig_ma = n_orig_pres = n_orig_empty = n_orig_api_err = 0
    merge_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for offset, col_name in enumerate(ENRICHMENT_ORDER):
        ws.cell(row=1, column=col_start + offset).value = col_name

    for idx, api_id in enumerate(api_id_list, start=2):
        if api_id is None or api_id == "":
            continue
        try:
            lid = int(api_id)
        except (TypeError, ValueError):
            continue
        match = df_enriched[df_enriched[config.API_LOAN_ID_COLUMN] == lid]
        if match.empty:
            continue
        d = match.iloc[0].to_dict()
        fee_ma_raw = d.get(config.ORIGINATION_FEE_COLUMN)
        existing_orig_cell = ws.cell(row=idx, column=col_start + orig_off).value
        orig_source: str | None = None
        api_err = d.get("Error")

        for offset, col_name in enumerate(ENRICHMENT_ORDER):
            val = d.get(col_name)
            if col_name == config.ORIGINATION_FEE_COLUMN:
                if _fee_is_missing_or_zero(val):
                    existing = ws.cell(row=idx, column=col_start + offset).value
                    if not _fee_is_missing_or_zero(existing):
                        val = existing
                        preserved += 1
                        orig_source = "preserved_master"
                    else:
                        orig_source = "empty_after_ma"
                else:
                    orig_source = "ma"
            ws.cell(row=idx, column=col_start + offset).value = (
                "" if val in (None, {}, []) else val
            )

        if api_err:
            orig_source = orig_source or "api_row_error"
        if not orig_source:
            orig_source = "unknown"

        if orig_source == "ma":
            n_orig_ma += 1
        elif orig_source == "preserved_master":
            n_orig_pres += 1
        elif orig_source == "empty_after_ma":
            n_orig_empty += 1
        elif orig_source == "api_row_error":
            n_orig_api_err += 1

        if dry_run:
            per_loan_rows.append(
                {
                    "loan_id": lid,
                    "origination_fee_from_ma_api": fee_ma_raw,
                    "origination_fee_master_cell_before_merge": existing_orig_cell,
                    "origination_fee_source_after_step2": orig_source,
                    "api_row_error": api_err or "",
                }
            )

    summary: dict = {
        "rows_on_sheet": len(api_id_list),
        "origination_fee_preserved": preserved,
        "enrich_loan_count": int(len(loan_ids)),
        "orig_from_ma": n_orig_ma,
        "orig_preserved_master": n_orig_pres,
        "orig_empty_after_ma": n_orig_empty,
        "orig_api_row_error": n_orig_api_err,
    }

    if dry_run:
        from .placement_fee_backfill import preview_backfill_on_worksheet

        bf_prev = preview_backfill_on_worksheet(ws)
        summary["backfill_preview_rows"] = bf_prev
        if per_loan_rows:
            p2 = config.LOGS_DIR / f"dryrun_step2_by_loan_{merge_stamp}.csv"
            pd.DataFrame(per_loan_rows).to_csv(p2, index=False, encoding="utf-8-sig")
            print(f"  [dry-run] per-loan enrich -> {p2}")
        if bf_prev:
            p3 = config.LOGS_DIR / f"dryrun_step3_backfill_preview_{merge_stamp}.csv"
            pd.DataFrame(bf_prev).to_csv(p3, index=False, encoding="utf-8-sig")
            print(f"  [dry-run] backfill preview (post-merge in memory) -> {p3}")
        print(
            f"  [dry-run] would save workbook; preserved_fee_cells={preserved}; "
            f"step3 backfill rows that would get Excel fee: {len(bf_prev)}"
        )
        wb.close()
    else:
        wb.save(config.EXCEL_TRACKER_PATH)
        _write_last_enricher_run_marker()
        print(
            f"  Workbook saved. origination_fee_preserved={preserved}; "
            f"last-run marker updated."
        )
        wb.close()
    return summary
