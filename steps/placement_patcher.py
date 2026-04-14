"""
Placement Fee Patcher — reads patcher workbook, updates MA when allowed.
"""
from __future__ import annotations

import hashlib
import math
import time
from datetime import datetime, timezone

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import config


def api_auth_token(action: str) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d-%H")
    raw = f"{config.MA_LENDER_ID}-{config.MA_API_KEY}-{action}-{ts}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def get_table_dataframe(xlsx_path: str, table_name: str) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = tbl = None
    for sheet in wb.worksheets:
        if table_name in sheet.tables:
            ws, tbl = sheet, sheet.tables[table_name]
            break
    if ws is None or tbl is None:
        wb.close()
        raise ValueError(f"Table '{table_name}' not found in workbook: {xlsx_path}")
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    rows = list(
        ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    wb.close()
    if not rows:
        return pd.DataFrame()
    headers = [("" if h is None else str(h).strip()) for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=headers)
    df = df.dropna(how="all").reset_index(drop=True)
    # Normalize duplicate/odd header keys pandas may introduce
    df.columns = [str(c).strip() for c in df.columns]
    return df


def diagnose_patcher_fees(df: pd.DataFrame) -> None:
    """Log rows where API Loan ID looks valid but Origination Fee is missing/zero (often stale formula cache)."""
    if df.empty:
        return
    lid_col = config.PATCHER_COL_LOAN_ID
    fee_col = config.PATCHER_COL_ORIG_FEE
    if lid_col not in df.columns or fee_col not in df.columns:
        print(
            f"  [patcher] WARN: expected columns {lid_col!r} / {fee_col!r}; "
            f"have: {list(df.columns)}"
        )
        return
    bad = 0
    samples = []
    for _, row in df.iterrows():
        lid = to_number(row.get(lid_col))
        fee = to_number(row.get(fee_col))
        if lid is None:
            continue
        if fee is None or fee <= 0:
            bad += 1
            if len(samples) < 5:
                raw = row.get(fee_col)
                samples.append((int(lid), raw, type(raw).__name__))
    if bad:
        print(
            f"  [patcher] WARN: {bad} row(s) have {fee_col!r} missing/zero/non-numeric "
            f"but {lid_col} set. Excel may need 'Recalculate workbook' then save "
            f"(openpyxl reads cached formula results with data_only=True)."
        )
        for s in samples:
            print(f"    sample loan_id={s[0]} raw_fee={s[1]!r} ({s[2]})")


def to_number(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def is_nonzero(val, tol: float = config.NONZERO_TOL) -> bool:
    try:
        return abs(float(val)) > tol
    except Exception:
        return False


def approx_equal(a, b, tol: float = config.VERIFY_TOL) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def _headers(action: str) -> dict:
    return {
        "Content-Type": "application/json",
        "ACCOUNT-ID": config.MA_LENDER_ID,
        "API-AUTH": api_auth_token(action),
    }


def loans_get(session: requests.Session, loan_id: int) -> dict:
    url = f"{config.MA_API_BASE_V1}/loans/get"
    try:
        r = session.post(
            url,
            headers=_headers("loans/get"),
            json={"loan_id": int(loan_id)},
            timeout=config.TIMEOUT_S,
        )
        return r.json()
    except Exception:
        return {}


def get_fee_and_title(session: requests.Session, loan_id: int) -> tuple:
    """
    Returns (placement_fee, loan_title, error_message).
    error_message is non-empty if loans/get did not return a usable result.
    """
    js = loans_get(session, loan_id)
    if not isinstance(js, dict) or not js.get("result"):
        excerpt = ""
        try:
            excerpt = str(js)[:400]
        except Exception:
            pass
        return None, None, excerpt or "loans/get returned no result (check MA_API_KEY / loan_id)"
    try:
        mort = js["result"]["mortgage"]
    except Exception:
        mort = {}
    fee = (mort.get("fees") or {}).get("placement_fee")
    loan_title = js.get("result", {}).get("loan_title")
    return fee, loan_title, ""


def post_update(session: requests.Session, base_url: str, loan_id: int, fee: float) -> dict:
    url = f"{base_url}/loans/update"
    payload = {
        "loan_id": int(loan_id),
        "mortgage": {"fees": {"placement_fee": round(float(fee), 2)}},
    }
    last_err = None
    for attempt in range(config.RETRIES_PER_ENDPOINT + 1):
        try:
            r = session.post(
                url,
                headers=_headers("loans/update"),
                json=payload,
                timeout=config.TIMEOUT_S,
            )
            text = r.text
            try:
                js = r.json()
            except Exception:
                js = None
            return {"status": r.status_code, "text": text, "json": js, "endpoint": url}
        except requests.RequestException as e:
            last_err = str(e)
            if attempt < config.RETRIES_PER_ENDPOINT:
                time.sleep(0.8 + attempt)
    return {"status": None, "text": last_err or "", "json": None, "endpoint": url}


def _is_old_core(res: dict) -> bool:
    t = (res.get("text") or "").lower()
    return "old core loans" in t or "not supported" in t


def _build_result(loan_id, loan_title, street, desired_fee, before, after, action, res) -> dict:
    return {
        "loan_id": loan_id,
        "loan_title": loan_title,
        "street": street,
        "desired_fee": round(desired_fee, 2),
        "existing_fee_before": before,
        "existing_fee_after": after,
        "action": action,
        "status": res.get("status"),
        "endpoint": res.get("endpoint"),
        "response_excerpt": (res.get("text") or "")[:500],
    }


def patch_one(
    session: requests.Session,
    loan_id: int,
    desired_fee: float,
    street: str,
    *,
    dry_run: bool,
) -> dict:
    existing_before, loan_title, get_err = get_fee_and_title(session, loan_id)
    if get_err:
        return {
            "loan_id": loan_id,
            "loan_title": loan_title,
            "street": street,
            "desired_fee": round(desired_fee, 2),
            "existing_fee_before": None,
            "existing_fee_after": None,
            "action": "ERROR_LOANS_GET",
            "status": None,
            "endpoint": None,
            "response_excerpt": get_err,
        }

    if config.SKIP_IF_EXISTING_NONZERO and existing_before is not None and is_nonzero(
        existing_before
    ):
        return {
            "loan_id": loan_id,
            "loan_title": loan_title,
            "street": street,
            "desired_fee": round(desired_fee, 2),
            "existing_fee_before": existing_before,
            "existing_fee_after": existing_before,
            "action": "SKIPPED_ALREADY_SET",
            "status": 200,
            "endpoint": None,
            "response_excerpt": "Existing non-zero placement_fee; left unchanged.",
        }

    if dry_run:
        # GET-only: we know skip vs "would try PATCH"; UPDATED vs OLD_CORE needs POST.
        action = "DRY_RUN_WOULD_ATTEMPT_PATCH"
        return {
            "loan_id": loan_id,
            "loan_title": loan_title,
            "street": street,
            "desired_fee": round(desired_fee, 2),
            "existing_fee_before": existing_before,
            "existing_fee_after": existing_before,
            "action": action,
            "status": None,
            "endpoint": None,
            "response_excerpt": (
                "[dry-run] no loans/update; POST would be required to know "
                "UPDATED vs NOT_SUPPORTED_OLD_CORE vs ERROR."
            ),
        }

    res_v1 = post_update(session, config.MA_API_BASE_V1, loan_id, desired_fee)
    if not _is_old_core(res_v1):
        time.sleep(config.VERIFY_SETTLE_WAIT)
        existing_after, _, _ = get_fee_and_title(session, loan_id)
        ok = existing_after is not None and approx_equal(existing_after, desired_fee)
        return _build_result(
            loan_id,
            loan_title,
            street,
            desired_fee,
            existing_before,
            existing_after,
            "UPDATED" if ok else "ERROR",
            res_v1,
        )

    res_v2 = post_update(session, config.MA_API_BASE_V2, loan_id, desired_fee)
    time.sleep(config.VERIFY_SETTLE_WAIT)
    existing_after, _, _ = get_fee_and_title(session, loan_id)
    if _is_old_core(res_v2):
        return _build_result(
            loan_id,
            loan_title,
            street,
            desired_fee,
            existing_before,
            existing_after,
            "NOT_SUPPORTED_OLD_CORE",
            res_v2,
        )

    ok = existing_after is not None and approx_equal(existing_after, desired_fee)
    return _build_result(
        loan_id,
        loan_title,
        street,
        desired_fee,
        existing_before,
        existing_after,
        "UPDATED" if ok else "ERROR",
        res_v2,
    )


def run_patcher(*, dry_run: bool = False) -> list[dict]:
    """
    Read patcher Excel, call MA API (or simulate if dry_run), write CSV to logs/.
    """
    df = get_table_dataframe(config.EXCEL_PATCHER_PATH, config.PATCHER_TABLE_NAME)
    diagnose_patcher_fees(df)
    work = []
    for _, row in df.iterrows():
        loan_id = to_number(row.get(config.PATCHER_COL_LOAN_ID))
        fee = to_number(row.get(config.PATCHER_COL_ORIG_FEE))
        street = row.get(config.PATCHER_COL_STREET)
        if loan_id is None or fee is None or fee <= 0:
            continue
        work.append((int(loan_id), round(float(fee), 2), street))

    if not work:
        return []

    session = requests.Session()
    results = []
    label = "[dry-run] " if dry_run else ""
    for i, (loan_id, fee, street) in enumerate(work, 1):
        rec = patch_one(session, loan_id, fee, street, dry_run=dry_run)
        results.append(rec)
        ep = rec["endpoint"] or ""
        ep_lbl = "v2" if "/v2/" in ep else ("v1" if ep else "-")
        print(
            f"  {label}[{i}/{len(work)}] loan={loan_id} -> {rec['action']} "
            f"(ep={ep_lbl}, before={rec['existing_fee_before']}, after={rec['existing_fee_after']})"
        )
        time.sleep(config.SLEEP_BETWEEN_CALLS)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "_dryrun" if dry_run else ""
    csv_path = config.LOGS_DIR / f"placement_patch_{stamp}{suffix}.csv"
    pd.DataFrame(results).to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  CSV saved -> {csv_path}")
    return results
