"""
Backfill Origination Fee on Master Tracker from Placement Fee Patching workbook
when the cell is still empty/zero after MA enrichment.
"""
from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

import config

from .placement_patcher import get_table_dataframe, to_number


def _fee_is_missing_or_zero(val) -> bool:
    if val is None or val == "":
        return True
    try:
        return abs(float(val)) <= config.NONZERO_TOL
    except (TypeError, ValueError):
        return True


def _find_header_column(ws, name: str) -> int | None:
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == name:
            return cell.column
    return None


def load_patcher_fees() -> dict[int, float]:
    df = get_table_dataframe(config.EXCEL_PATCHER_PATH, config.PATCHER_TABLE_NAME)
    out: dict[int, float] = {}
    for _, row in df.iterrows():
        lid = to_number(row.get(config.PATCHER_COL_LOAN_ID))
        fee = to_number(row.get(config.PATCHER_COL_ORIG_FEE))
        if lid is None or fee is None or fee <= 0:
            continue
        out[int(lid)] = round(float(fee), 2)
    return out


def preview_backfill_on_worksheet(ws) -> list[dict]:
    """
    After enricher merged into ws in memory (dry-run), list rows that step-3 would fill
    from the Placement Fee Patching workbook.
    """
    fees_by_loan = load_patcher_fees()
    col_id = _find_header_column(ws, config.API_LOAN_ID_COLUMN)
    col_fee = _find_header_column(ws, config.ORIGINATION_FEE_COLUMN)
    if col_id is None or col_fee is None:
        return []
    out: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=col_id).value
        if raw_id is None or raw_id == "":
            continue
        try:
            lid = int(raw_id)
        except (TypeError, ValueError):
            continue
        current = ws.cell(row=row_idx, column=col_fee).value
        if not _fee_is_missing_or_zero(current):
            continue
        desired = fees_by_loan.get(lid)
        if desired is None:
            continue
        out.append(
            {
                "loan_id": lid,
                "master_excel_row": row_idx,
                "fee_would_set_from_patcher_excel": desired,
            }
        )
    return out


def run_backfill(*, dry_run: bool = False) -> dict:
    fees_by_loan = load_patcher_fees()
    wb = load_workbook(config.EXCEL_TRACKER_PATH)
    ws = wb[config.ENRICHER_SHEET_NAME]

    col_id = _find_header_column(ws, config.API_LOAN_ID_COLUMN)
    col_fee = _find_header_column(ws, config.ORIGINATION_FEE_COLUMN)
    if col_id is None:
        wb.close()
        raise ValueError(
            f"Column '{config.API_LOAN_ID_COLUMN}' not found in row 1 of sheet "
            f"'{config.ENRICHER_SHEET_NAME}'."
        )
    if col_fee is None:
        wb.close()
        raise ValueError(
            f"Column '{config.ORIGINATION_FEE_COLUMN}' not found in row 1. "
            "Run enricher once so headers exist, or add the column."
        )

    filled = 0
    skipped_has_value = 0
    skipped_no_patcher_row = 0

    for row_idx in range(2, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=col_id).value
        if raw_id is None or raw_id == "":
            continue
        try:
            lid = int(raw_id)
        except (TypeError, ValueError):
            continue
        current = ws.cell(row=row_idx, column=col_fee).value
        if not _fee_is_missing_or_zero(current):
            skipped_has_value += 1
            continue
        desired = fees_by_loan.get(lid)
        if desired is None:
            skipped_no_patcher_row += 1
            continue
        if dry_run:
            print(f"  [dry-run] row {row_idx} loan {lid}: would set fee -> {desired}")
        else:
            ws.cell(row=row_idx, column=col_fee).value = desired
        filled += 1

    summary = {
        "filled_from_patcher_excel": filled,
        "skipped_row_already_has_fee": skipped_has_value,
        "skipped_no_fee_in_patcher_table": skipped_no_patcher_row,
        "patcher_table_loans": len(fees_by_loan),
    }

    if dry_run:
        print(f"  [dry-run] backfill summary: {summary}")
        wb.close()
    else:
        wb.save(config.EXCEL_TRACKER_PATH)
        print(f"  Backfill saved. summary: {summary}")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(config.LOGS_DIR / f"backfill_summary_{stamp}.txt", "w", encoding="utf-8") as f:
        for k, v in summary.items():
            f.write(f"{k}={v}\n")

    return summary
